def safe_get_widget_value(widget, widget_name="unknown"):
    """Safely get value from any widget type - reusable function"""
    try:
        if widget is None:
            print(f"[WARNING] Widget {widget_name} is None")
            return ""
        
        # For tkcalendar DateEntry
        if hasattr(widget, 'get_date'):
            try:
                date_obj = widget.get_date()
                return date_obj.strftime("%d-%m-%Y")
            except Exception as e:
                print(f"[WARNING] Error getting date from {widget_name}: {e}")
                # Fallback to regular get()
                return str(widget.get()) if hasattr(widget, 'get') else ""
        
        # For regular Entry widgets
        elif hasattr(widget, 'get'):
            # Check if get method requires arguments
            import inspect
            sig = inspect.signature(widget.get)
            if len(sig.parameters) == 0:
                return str(widget.get())
            else:
                # For widgets like Listbox that need selection
                try:
                    selection = widget.curselection()
                    if selection:
                        return str(widget.get(selection[0]))
                    else:
                        return ""
                except:
                    return ""
        
        # For Text widgets
        elif hasattr(widget, 'get') and hasattr(widget, 'insert'):
            return str(widget.get("1.0", "end")).strip()
        
        else:
            print(f"[WARNING] Unknown widget type for {widget_name}: {type(widget)}")
            return ""
            
    except Exception as e:
        print(f"[ERROR] Error getting value from {widget_name}: {e}")
        import traceback
        traceback.print_exc()
        return ""

def safe_get_text_widget_value(widget, widget_name="unknown"):
    """Safely get value from Text widgets"""
    try:
        if widget is None:
            print(f"[WARNING] Text widget {widget_name} is None")
            return ""
        
        if hasattr(widget, 'get'):
            return widget.get("1.0", "end").strip()
        else:
            print(f"[WARNING] Text widget {widget_name} doesn't have get method")
            return ""
            
    except Exception as e:
        print(f"[ERROR] Error getting text value from {widget_name}: {e}")
        import traceback
        traceback.print_exc()
        return ""

def collect_form_data_safely(self):
    """Safely collect all form data with proper error handling"""
    try:
        # Start with vars data (IntVar, StringVar, etc.)
        data = {}
        if hasattr(self, 'vars') and self.vars:
            for key, var in self.vars.items():
                try:
                    data[key] = var.get()
                except Exception as e:
                    print(f"[WARNING] Error getting value from var {key}: {e}")
                    data[key] = 0 if 'IntVar' in str(type(var)) else ""
        
        # Get data from form input widgets
        if hasattr(self, 'form_input_widgets') and self.form_input_widgets:
            # Text widgets (multiline)
            text_fields = ["perihal", "asal_surat", "ditujukan", "bicarakan_dengan", "teruskan_kepada"]
            for field in text_fields:
                if field in self.form_input_widgets:
                    data[field] = safe_get_text_widget_value(
                        self.form_input_widgets[field], 
                        field
                    )
                else:
                    data[field] = ""
            
            # Date widget for tgl_surat
            if "tgl_surat" in self.form_input_widgets:
                data["tgl_surat"] = safe_get_widget_value(
                    self.form_input_widgets["tgl_surat"], 
                    "tgl_surat"
                )
            else:
                data["tgl_surat"] = ""
        
        # Handle special date entries
        if hasattr(self, 'tgl_terima_entry'):
            data["tgl_terima"] = safe_get_widget_value(
                self.tgl_terima_entry, 
                "tgl_terima_entry"
            )
        else:
            data["tgl_terima"] = ""
        
        if hasattr(self, 'harap_selesai_tgl_entry'):
            data["harap_selesai_tgl"] = safe_get_widget_value(
                self.harap_selesai_tgl_entry, 
                "harap_selesai_tgl_entry"
            )
        else:
            data["harap_selesai_tgl"] = ""
        
        # Ensure required fields have default values
        data["indeks"] = data.get("indeks", "")
        data["rahasia"] = data.get("rahasia", 0)
        data["penting"] = data.get("penting", 0)
        data["segera"] = data.get("segera", 0)
        
        # Collect instructions safely
        if hasattr(self, "instruksi_table") and hasattr(self.instruksi_table, "get_data"):
            try:
                data["isi_instruksi"] = self.instruksi_table.get_data()
            except Exception as e:
                print(f"[WARNING] Error getting instruction data: {e}")
                data["isi_instruksi"] = []
        else:
            data["isi_instruksi"] = []
        
        return data
        
    except Exception as e:
        print(f"[ERROR] Error collecting form data: {e}")
        import traceback
        traceback.print_exc()
        return {}

def save_to_pdf(self):
    import threading
    from tkinter import filedialog
    import time, traceback, os, tempfile
    from pdf_output import save_form_to_pdf, merge_pdfs
    from sheet_logic import upload_to_sheet
    from disposisi_app.views.components.loading_screen import loading_manager, LoadingMessageBox
    
    def do_save():
        try:
            # Show loading screen
            loading_manager.show_loading(self, "Saving to PDF...", True)
            
            # Progress simulation
            for i in range(1, 101):
                time.sleep(0.02)
                loading_manager.update_progress(i)
            
            filepath = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF Documents", "*.pdf"), ("All Files", "*.*")]
            )
            if not filepath:
                loading_manager.hide_loading()
                return
            
            # Safely collect form data
            data = collect_form_data_safely(self)
            
            # Validate required fields
            if not str(data.get("no_surat", "")).strip():
                LoadingMessageBox.showerror("Validasi", "No. Surat tidak boleh kosong!", parent=self)
                loading_manager.hide_loading()
                return
            
            # Check uniqueness
            try:
                from disposisi_app.views.components.validation import is_no_surat_unique
                from google_sheets_connect import get_sheets_service, SHEET_ID
                if not is_no_surat_unique(data.get("no_surat", ""), get_sheets_service, SHEET_ID):
                    LoadingMessageBox.showerror("Validasi", "No. Surat sudah ada di sheet, harus unik!", parent=self)
                    loading_manager.hide_loading()
                    return
            except Exception as e:
                print(f"[WARNING] Error checking uniqueness: {e}")
                # Continue anyway if validation fails
            
            # Create PDF
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
                temp_pdf_path = temp_pdf.name
            
            save_form_to_pdf(temp_pdf_path, data)
            
            # Merge with attachments if they exist
            pdf_list = [temp_pdf_path]
            if hasattr(self, 'pdf_attachments') and self.pdf_attachments:
                pdf_list.extend(list(self.pdf_attachments))
            
            merge_pdfs(pdf_list, filepath)
            
            # Clean up temp file
            try:
                os.remove(temp_pdf_path)
            except Exception as e:
                print(f"[WARNING] Error removing temp file: {e}")
            
            # Upload to Google Sheets
            try:
                upload_to_sheet(self, call_from_pdf=True, data_override=data)
            except Exception as e:
                traceback.print_exc()
                LoadingMessageBox.showerror("Google Sheets", f"Gagal upload ke Google Sheets: {e}", parent=self)
            
            # Clean up and show success
            loading_manager.hide_loading()
            LoadingMessageBox.showinfo("Sukses", f"Formulir disposisi dan lampiran telah disimpan di:\n{filepath}", parent=self)
            
            # Reset form state
            self.edit_mode = False
            if hasattr(self, 'pdf_attachments'):
                self.pdf_attachments = []
            if hasattr(self, 'refresh_pdf_attachments'):
                try:
                    self.refresh_pdf_attachments(self._form_main_frame)
                except Exception as e:
                    print(f"[WARNING] Error refreshing PDF attachments: {e}")
            
        except Exception as e:
            traceback.print_exc()
            LoadingMessageBox.showerror("Error", f"Terjadi error: {e}", parent=self)
        finally:
            loading_manager.hide_loading()
    
    threading.Thread(target=do_save, daemon=True).start()

def save_to_sheet(self):
    import threading
    import time, traceback
    from sheet_logic import upload_to_sheet
    from disposisi_app.views.components.loading_screen import loading_manager, LoadingMessageBox
    
    def do_save():
        try:
            # Show loading screen
            loading_manager.show_loading(self, "Saving to Sheet...", True)
            
            # Progress simulation
            for i in range(1, 101):
                time.sleep(0.01)
                loading_manager.update_progress(i)
            
            # Safely collect form data
            data = collect_form_data_safely(self)
            
            # Validate required fields
            if not str(data.get("no_surat", "")).strip():
                LoadingMessageBox.showerror("Validasi", "No. Surat tidak boleh kosong!", parent=self)
                loading_manager.hide_loading()
                return
            
            # Check uniqueness
            try:
                from disposisi_app.views.components.validation import is_no_surat_unique
                from google_sheets_connect import get_sheets_service, SHEET_ID
                if not is_no_surat_unique(data.get("no_surat", ""), get_sheets_service, SHEET_ID):
                    LoadingMessageBox.showerror("Validasi", "No. Surat sudah ada di sheet, harus unik!", parent=self)
                    loading_manager.hide_loading()
                    return
            except Exception as e:
                print(f"[WARNING] Error checking uniqueness: {e}")
                LoadingMessageBox.showerror("Error", f"Error validasi: {e}", parent=self)
                loading_manager.hide_loading()
                return
            
            # Upload to sheet
            upload_to_sheet(self, call_from_pdf=False)
            
            loading_manager.hide_loading()
                
        except Exception as e:
            traceback.print_exc()
            LoadingMessageBox.showerror("Error", f"Terjadi error: {e}", parent=self)
        finally:
            loading_manager.hide_loading()
    
    threading.Thread(target=do_save, daemon=True).start()

def safe_get_filter_value(widget, widget_name="unknown", default=""):
    """Safely get value from filter widgets"""
    try:
        if widget is None:
            return default
        
        if hasattr(widget, 'get'):
            return str(widget.get())
        else:
            return default
            
    except Exception as e:
        print(f"[WARNING] Error getting filter value from {widget_name}: {e}")
        return default

def export_excel_advanced(self, filepath):
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter
    from tkinter import messagebox
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.worksheets[0]
        if ws is None:
            messagebox.showerror("Export Excel", "Worksheet tidak dapat dibuat.")
            return
        
        ws.title = "Log Disposisi"
        
        # Header setup
        ws.merge_cells('A1:AH1')
        ws['A1'] = 'LAPORAN DISPOSISI'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        import datetime
        today_str = datetime.datetime.now().strftime('%d-%m-%Y')
        ws.merge_cells('A2:AH2')
        ws['A2'] = f'terakhir di update: {today_str}'
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Filter information - safely get values
        filter_info = []
        
        disposisi_val = safe_get_filter_value(
            getattr(self, 'filter_val', None), 
            'filter_val'
        )
        if disposisi_val:
            filter_info.append(f"Disposisi ke: {disposisi_val}")
        
        bulan = safe_get_filter_value(
            getattr(self, 'month_var', None), 
            'month_var'
        )
        if bulan:
            filter_info.append(f"Bulan: {bulan}")
        
        tahun = safe_get_filter_value(
            getattr(self, 'year_var', None), 
            'year_var'
        )
        if tahun:
            filter_info.append(f"Tahun: {tahun}")
        
        search_col = safe_get_filter_value(
            getattr(self, 'search_col', None), 
            'search_col'
        )
        search_val = safe_get_filter_value(
            getattr(self, 'search_entry', None), 
            'search_entry'
        )
        if search_col and search_val:
            filter_info.append(f"Pencarian: {search_col} = {search_val}")
        
        ws.merge_cells('A3:AH3')
        ws['A3'] = ' | '.join(filter_info) if filter_info else 'Semua Data'
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers
        header_row3 = [
            "No. Agenda", "No. Surat", "Tgl. Surat", "Perihal", "Asal Surat", "Ditujukan", 
            "Klasifikasi", "Disposisi kepada", "Untuk Di :", "Selesai Tgl.", "Kode Klasifikasi", 
            "Tgl. Penerimaan", "Indeks", "Bicarakan dengan", "Teruskan kepada", "Harap Selesai Tanggal",
            "Direktur Utama", "", "Direktur Keuangan", "", "Direktur Teknik", "", 
            "GM Keuangan & Administrasi", "", "GM Operasional & Pemeliharaan", "", 
            "Manager Pemeliharaan", "", "Manager Operasional", "", "Manager Administrasi", "", "Manager Keuangan", ""
        ]
        ws.append(header_row3)
        
        header_row4 = ["" for _ in range(28)]
        positions = [16, 18, 20, 22, 24, 26]  # Instruction columns for existing positions
        for pos in positions:
            if pos < 28:  # Ensure we don't exceed the range
                header_row4[pos] = "Instruksi"
                if pos + 1 < 28:  # Ensure we don't exceed the range
                    header_row4[pos + 1] = "Tanggal"
        ws.append(header_row4)
        
        # Enhanced header with 34 columns (A-AH) matching Google Sheets structure
        ENHANCED_HEADER = [
            "No. Agenda", "No. Surat", "Tgl. Surat", "Perihal", "Asal Surat", "Ditujukan",
            "Klasifikasi", "Disposisi kepada", "Untuk Di :", "Selesai Tgl.", "Kode Klasifikasi", 
            "Tgl. Penerimaan", "Indeks", "Bicarakan dengan", "Teruskan kepada", "Harap Selesai Tanggal",
            "Direktur Utama Instruksi", "Direktur Utama Tanggal", "Direktur Keuangan Instruksi", "Direktur Keuangan Tanggal",
            "Direktur Teknik Instruksi", "Direktur Teknik Tanggal", "GM Keuangan & Administrasi Instruksi", "GM Keuangan & Administrasi Tanggal",
            "GM Operasional & Pemeliharaan Instruksi", "GM Operasional & Pemeliharaan Tanggal", 
            "Manager Pemeliharaan Instruksi", "Manager Pemeliharaan Tanggal",
            "Manager Operasional Instruksi", "Manager Operasional Tanggal",
            "Manager Administrasi Instruksi", "Manager Administrasi Tanggal",
            "Manager Keuangan Instruksi", "Manager Keuangan Tanggal"
        ]
        
        # Add data rows safely
        if hasattr(self, 'filtered_data') and self.filtered_data:
            for row in self.filtered_data:
                try:
                    row_data = []
                    for col in ENHANCED_HEADER:
                        cell_value = row.get(col, "")
                        # Handle None values
                        if cell_value is None:
                            cell_value = ""
                        row_data.append(str(cell_value))
                    ws.append(row_data)
                except Exception as e:
                    print(f"[WARNING] Error adding row to Excel: {e}")
                    # Add empty row to maintain structure
                    ws.append(["" for _ in range(len(ENHANCED_HEADER))])
        else:
            # Add a placeholder row if no data
            ws.append(["Tidak ada data" for _ in range(len(ENHANCED_HEADER))])
        
        # Apply formatting
        thin = Side(border_style="thin", color="000000")
        try:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=28):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        except Exception as e:
            print(f"[WARNING] Error applying cell formatting: {e}")
        
        # Adjust column widths
        try:
            for i, col in enumerate(ENHANCED_HEADER, 1):
                try:
                    max_length = 10  # Default width
                    for r in range(1, ws.max_row + 1):
                        cell_value = ws.cell(row=r, column=i).value
                        if cell_value is not None:
                            length = len(str(cell_value))
                            if length > max_length:
                                max_length = length
                    
                    ws.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 40)
                except Exception as e:
                    print(f"[WARNING] Error adjusting column {i} width: {e}")
                    ws.column_dimensions[get_column_letter(i)].width = 15
        except Exception as e:
            print(f"[WARNING] Error in column width adjustment: {e}")
        
        # Save file
        wb.save(filepath)
        messagebox.showinfo("Export Excel", f"Data berhasil diekspor ke {filepath}")
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        messagebox.showerror("Export Excel", f"Gagal ekspor: {e}")
def send_email_with_disposisi(self, selected_positions):
    """
    ENHANCED VERSION: Mengirimkan email disposisi ke posisi yang dipilih dengan dukungan Senior Officer.
    Ketika Manager dipilih, otomatis akan muncul pilihan 2 Senior Officer terkait.
    """
    import os
    import tempfile
    import traceback
    from tkinter import messagebox
    from datetime import datetime
    
    # Import modul email yang sudah diperbaiki
    try:
        from email_sender.send_email import EmailSender
        from email_sender.template_handler import render_email_template
        from pdf_output import save_form_to_pdf, merge_pdfs
    except ImportError as e:
        messagebox.showerror("Import Error", f"Gagal import modul email: {e}")
        return

    print(f"[DEBUG] Selected positions: {selected_positions}")
    self.update_status("Menyiapkan pengiriman email...")
    
    # Collect form data
    data = collect_form_data_safely(self)
    if not data.get("no_surat", "").strip():
        messagebox.showerror("Validation Error", "No. Surat tidak boleh kosong untuk mengirim email.")
        return

    # Generate PDF attachment
    temp_pdf_path = None
    final_pdf_path = None
    
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
            temp_pdf_path = temp_pdf.name
        
        save_form_to_pdf(temp_pdf_path, data)
        
        final_pdf_path = temp_pdf_path
        if hasattr(self, 'pdf_attachments') and self.pdf_attachments:
            merged_pdf_path = os.path.join(tempfile.gettempdir(), f"merged_{os.path.basename(temp_pdf_path)}")
            pdf_list = [temp_pdf_path] + self.pdf_attachments
            merge_pdfs(pdf_list, merged_pdf_path)
            final_pdf_path = merged_pdf_path

    except Exception as e:
        messagebox.showerror("PDF Generation Error", f"Gagal membuat PDF untuk email: {e}")
        traceback.print_exc()
        return
    
    # Initialize EmailSender
    try:
        email_sender = EmailSender()
        
        if not email_sender.sheets_service:
            error_msg = ("Google Sheets tidak dapat diakses.\n\n"
                        "Pastikan:\n"
                        "1. File credentials/credentials.json tersedia\n"
                        "2. Sheet admin dengan email sudah dibuat\n"
                        "3. ID sheet admin benar di config.py")
            messagebox.showerror("Email Configuration Error", error_msg)
            return
                
    except Exception as e:
        error_msg = f"Gagal menginisialisasi email sender: {str(e)}"
        messagebox.showerror("Email Sender Error", error_msg)
        traceback.print_exc()
        return
    
    # ENHANCED: Ambil email untuk posisi yang dipilih (termasuk senior officers)
    self.update_status("Mencari alamat email dari database admin...")
    
    recipient_emails = []
    failed_lookups = []
    successful_lookups = []
    
    print(f"[DEBUG] Looking up emails for positions: {selected_positions}")
    
    for position in selected_positions:
        print(f"[DEBUG] Looking up email for: {position}")
        email, msg = email_sender.get_recipient_email(position)
        print(f"[DEBUG] Result for {position}: email={email}, msg={msg}")
        
        if email and email != position:
            # Validasi format email
            if '@' in email and '.' in email.split('@')[-1]:
                recipient_emails.append(email)
                successful_lookups.append(f"{position}: {email}")
                print(f"[DEBUG] ✓ Valid email found: {position} -> {email}")
            else:
                failed_lookups.append(f"{position} (invalid email format: {email})")
                print(f"[DEBUG] ✗ Invalid email format: {position} -> {email}")
        else:
            failed_lookups.append(f"{position} ({msg})")
            print(f"[DEBUG] ✗ No email found: {position} -> {msg}")
    
    print(f"[DEBUG] Final recipient emails: {recipient_emails}")
    print(f"[DEBUG] Failed lookups: {failed_lookups}")
    
    # Show warnings for failed lookups
    if failed_lookups:
        # Gunakan singkatan untuk manager dalam warning message
        abbreviation_map = {
            "Manager Pemeliharaan": "pml",
            "Manager Operasional": "ops",
            "Manager Administrasi": "adm",
            "Manager Keuangan": "keu"
        }
        
        # Konversi failed lookups ke singkatan untuk display
        display_failed_lookups = []
        for lookup in failed_lookups:
            for full_name, abbrev in abbreviation_map.items():
                if full_name in lookup:
                    lookup = lookup.replace(full_name, f"Manager {abbrev}")
                    break
            display_failed_lookups.append(lookup)
        
        warning_msg = ("Tidak dapat menemukan alamat email yang valid untuk posisi berikut:\n" + 
                      "\n".join([f"• {lookup}" for lookup in display_failed_lookups]))
        
        if len(failed_lookups) == len(selected_positions):
            # All positions failed
            full_error = (warning_msg + 
                         "\n\nPastikan admin sudah mengisi email yang valid untuk posisi-posisi tersebut "
                         "di spreadsheet admin.\n\n"
                         "Format email harus: user@domain.com")
            messagebox.showerror("No Valid Emails Found", full_error)
            return
        else:
            # Some positions failed - show warning but continue
            result = messagebox.askyesno(
                "Some Emails Missing", 
                warning_msg + "\n\nLanjutkan mengirim ke email yang valid?", 
                parent=self
            )
            if not result:
                return
    
    if not recipient_emails:
        messagebox.showerror("No Recipients", "Tidak ada alamat email yang valid untuk dikirimi.")
        return
    
    # ENHANCED: Prepare email template data with senior officer support
    template_data = {
        'nomor_surat': data.get('no_surat', 'N/A'),
        'nama_pengirim': data.get('asal_surat', 'N/A'),
        'perihal': data.get('perihal', 'N/A'),
        'tanggal': datetime.now().strftime('%d %B %Y'),
        'klasifikasi': [],
        'instruksi_list': [],
        'tahun': datetime.now().year
    }
    
    # Tambahkan informasi disposisi kepada dengan format abbreviation
    try:
        disposisi_labels = self.get_disposisi_labels_with_abbreviation()
        template_data['disposisi_kepada'] = disposisi_labels
    except Exception as e:
        print(f"[WARNING] Error getting disposisi labels: {e}")
        template_data['disposisi_kepada'] = []
    
    # ENHANCED: Add selected recipients info to template
    recipient_list = []
    for position in selected_positions:
        # Use abbreviation for display in email
        abbreviation_map = {
            "Manager Pemeliharaan": "pml",
            "Manager Operasional": "ops", 
            "Manager Administrasi": "adm",
            "Manager Keuangan": "keu"
        }
        
        display_position = abbreviation_map.get(position, position)
        if display_position in ["pml", "ops", "adm", "keu"]:
            display_position = f"Manager {display_position}"
        
        recipient_list.append(display_position)
    
    template_data['selected_recipients'] = recipient_list
    
    # Add classification
    if data.get('rahasia', 0):
        template_data['klasifikasi'].append("RAHASIA")
    if data.get('penting', 0):
        template_data['klasifikasi'].append("PENTING")
    if data.get('segera', 0):
        template_data['klasifikasi'].append("SEGERA")
    
    # Add instructions from checkboxes
    instruksi_mapping = [
        ("ketahui_file", "Ketahui & File"),
        ("proses_selesai", "Proses Selesai"),
        ("teliti_pendapat", "Teliti & Pendapat"),
        ("buatkan_resume", "Buatkan Resume"),
        ("edarkan", "Edarkan"),
        ("sesuai_disposisi", "Sesuai Disposisi"),
        ("bicarakan_saya", "Bicarakan dengan Saya")
    ]
    
    for key, label in instruksi_mapping:
        if data.get(key, 0):
            template_data['instruksi_list'].append(label)
    
    # Add detailed instructions from table
    if 'isi_instruksi' in data and data['isi_instruksi']:
        for instr in data['isi_instruksi']:
            if instr.get('instruksi', '').strip():
                posisi = instr.get('posisi', '')
                instruksi_text = instr.get('instruksi', '')
                tanggal = instr.get('tanggal', '')
                
                # Gunakan singkatan untuk manager dalam instruksi
                abbreviation_map = {
                    "Manager Pemeliharaan": "pml",
                    "Manager Operasional": "ops",
                    "Manager Administrasi": "adm",
                    "Manager Keuangan": "keu"
                }
                
                # Konversi posisi ke singkatan jika ada
                display_posisi = abbreviation_map.get(posisi, posisi)
                if display_posisi in ["pml", "ops", "adm", "keu"]:
                    display_posisi = f"Manager {display_posisi}"
                
                instr_line = f"{display_posisi}: {instruksi_text}"
                if tanggal:
                    instr_line += f" (Tanggal: {tanggal})"
                template_data['instruksi_list'].append(instr_line)
    
    # Add additional instructions
    if data.get("bicarakan_dengan", "").strip():
        template_data['instruksi_list'].append(f"Bicarakan dengan: {data['bicarakan_dengan']}")
    
    if data.get("teruskan_kepada", "").strip():
        template_data['instruksi_list'].append(f"Teruskan kepada: {data['teruskan_kepada']}")
    
    if data.get("harap_selesai_tgl", "").strip():
        template_data['instruksi_list'].append(f"Harap diselesaikan tanggal: {data['harap_selesai_tgl']}")
    
    # Render email template and send
    try:
        html_content = render_email_template(template_data)
        subject = f"Disposisi Surat: {data.get('perihal', 'N/A')}"
        
        self.update_status(f"Mengirim email ke {len(recipient_emails)} penerima...")
        print(f"[DEBUG] Sending email to: {recipient_emails}")
        print(f"[DEBUG] Subject: {subject}")
        print(f"[DEBUG] PDF attachment: {final_pdf_path}")
        
        # ENHANCED: Use method with PDF attachment
        success, message = email_sender.send_disposisi_email(
            recipient_emails, 
            subject, 
            html_content,
            pdf_attachment=final_pdf_path
        )
        
        print(f"[DEBUG] Email send result: success={success}, message={message}")
        
        # Show results
        if success:
            self.update_status("Email berhasil dikirim!")
            success_msg = f"Email berhasil dikirim ke:\n{chr(10).join([f'• {email}' for email in recipient_emails])}"
            
            if successful_lookups:
                # Gunakan singkatan untuk manager dalam successful lookups
                abbreviation_map = {
                    "Manager Pemeliharaan": "pml",
                    "Manager Operasional": "ops",
                    "Manager Administrasi": "adm",
                    "Manager Keuangan": "keu"
                }
                
                # Konversi successful lookups ke singkatan untuk display
                display_successful_lookups = []
                for lookup in successful_lookups:
                    for full_name, abbrev in abbreviation_map.items():
                        if full_name in lookup:
                            lookup = lookup.replace(full_name, f"Manager {abbrev}")
                            break
                    display_successful_lookups.append(lookup)
                
                success_msg += f"\n\nDetail penerima:\n{chr(10).join([f'• {lookup}' for lookup in display_successful_lookups])}"
                
                # ENHANCED: Show breakdown of recipient types
                managers = [p for p in selected_positions if p.startswith("Manager")]
                senior_officers = [p for p in selected_positions if p.startswith("Senior Officer")]
                others = [p for p in selected_positions if not p.startswith("Manager") and not p.startswith("Senior Officer")]
                
                if managers or senior_officers or others:
                    success_msg += "\n\nBreakdown penerima:"
                    if others:
                        success_msg += f"\n• Directors & GM: {len(others)}"
                    if managers:
                        success_msg += f"\n• Managers: {len(managers)}"
                    if senior_officers:
                        success_msg += f"\n• Senior Officers: {len(senior_officers)}"
            
            messagebox.showinfo("Email Sent Successfully", success_msg, parent=self)
        else:
            self.update_status("Gagal mengirim email.")
            messagebox.showerror("Email Send Failed", f"Gagal mengirim email:\n{message}", parent=self)
            
    except Exception as e:
        self.update_status("Error saat mengirim email.")
        messagebox.showerror("Email Error", f"Terjadi kesalahan saat mengirim email: {e}")
        traceback.print_exc()
    
    finally:
        # Clean up temporary files
        try:
            if temp_pdf_path and os.path.exists(temp_pdf_path):
                os.remove(temp_pdf_path)
                print(f"[DEBUG] Cleaned up temp PDF: {temp_pdf_path}")
            if final_pdf_path and final_pdf_path != temp_pdf_path and os.path.exists(final_pdf_path):
                os.remove(final_pdf_path)
                print(f"[DEBUG] Cleaned up final PDF: {final_pdf_path}")
        except Exception as e:
            print(f"[WARNING] Could not remove temporary files: {e}")